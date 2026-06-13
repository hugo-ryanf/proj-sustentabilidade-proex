"""Exportação dos projetos coletados para uma planilha Excel, uma aba por centro."""

import sqlite3

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment

from scraper.config import XLSX_PATH, COLUNAS_EXCEL
from scraper.database import carregar_projetos_db

CABECALHOS = ["Código", "Título", "Coordenador", "Centro", "Unidade", "Área Temática"]


def exportar_excel(conn: sqlite3.Connection):
    todos_projetos = carregar_projetos_db(conn)

    if not todos_projetos:
        print("Nenhum projeto relevante encontrado.")
        return

    header_fill  = PatternFill("solid", fgColor="1F4E79")
    header_font  = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")

    with pd.ExcelWriter(XLSX_PATH, engine="openpyxl") as writer:
        for centro_nome, projetos in todos_projetos.items():
            df = pd.DataFrame(projetos, columns=COLUNAS_EXCEL)
            df.columns = CABECALHOS
            df.to_excel(writer, sheet_name=centro_nome[:31], index=False)

            ws = writer.sheets[centro_nome[:31]]

            for cell in ws[1]:
                cell.fill      = header_fill
                cell.font      = header_font
                cell.alignment = header_align

            for col in ws.columns:
                max_len = max((len(str(c.value)) for c in col if c.value), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

            ws.freeze_panes = "A2"

    total = sum(len(p) for p in todos_projetos.values())
    print(f"\n{total} projeto(s) relevante(s) salvos em '{XLSX_PATH}'.")
